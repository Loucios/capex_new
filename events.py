from collections import defaultdict
from copy import copy

from base_datas import (NDS, ChapterDirect, CTPCostIndicator, DeflatorIndex,
                        NetworkCostIndicator, NetworkEvent,
                        SourceCostIndicator, SourceEvent, Stage, Terms,
                        TfuUnitCost, TSOList)
from openpyxl import load_workbook
from summary_datas import SummaryTable
from titles import Titles


class BaseMixin:
    # Наименование таблицы в Excel -> Наименование датакласса
    classes = {
        Titles.stages: Stage,
        Titles.deflator_indices: DeflatorIndex,
        Titles.event_years: Terms,
        Titles.nds: NDS,
        Titles.source_unit_costs: SourceCostIndicator,
        Titles.tfu_unit_costs: TfuUnitCost,
        Titles.source_events: SourceEvent,
        Titles.network_unit_costs: NetworkCostIndicator,
        Titles.ctp_unit_costs: CTPCostIndicator,
        Titles.tso_list: TSOList,
        Titles.network_events: NetworkEvent,
        Titles.chapter_7_directions: ChapterDirect,
        Titles.chapter_8_directions: ChapterDirect,
        Titles.source_ch12_directions: ChapterDirect,
        Titles.network_ch12_directions: ChapterDirect,
    }

    # Группа мероприятий по структуре направлений
    group_event = {
        Titles.chapter_7_directions: 'source_events',
        Titles.chapter_8_directions: 'network_events',
        Titles.source_ch12_directions: 'source_events',
        Titles.network_ch12_directions: 'network_events',
    }

    # Наименование атрибута
    atributes = {
        Titles.chapter_7_directions: 'chapter_7_directions',
        Titles.chapter_8_directions: 'chapter_8_directions',
        Titles.source_ch12_directions: 'source_ch12_directions',
        Titles.network_ch12_directions: 'network_ch12_directions',
    }


class Events(BaseMixin):
    def __init__(self, filename: str) -> None:
        self.wb, self.table_names = self._get_wb_and_table_names(filename)

        # import secondary tables
        source_unit_costs = self._get_table(Titles.source_unit_costs)
        tfu_unit_costs = self._get_table(Titles.tfu_unit_costs)
        network_unit_costs = self._get_table(Titles.network_unit_costs)
        ctp_unit_costs = self._get_table(Titles.ctp_unit_costs)
        stages = self._get_table(Titles.stages)
        deflator_indices = self._get_table(Titles.deflator_indices)
        self.terms = self._get_table(Titles.event_years)

        cum_deflator = 1
        for index in deflator_indices:
            if index.year >= self.terms[0].cost_year:
                cum_deflator *= index.index
                index.cumulative = cum_deflator

        nds = self._get_table(Titles.nds)

        # import and saving main tables
        base_params = (stages, deflator_indices, self.terms, nds)
        source_params = (source_unit_costs, tfu_unit_costs)
        network_params = (network_unit_costs, ctp_unit_costs)
        self.source_events, self.source_total = self._get_table(
                                                        Titles.source_events,
                                                        *base_params,
                                                        *source_params)
        self.network_events, self.network_total = self._get_table(
                                                        Titles.network_events,
                                                        *base_params,
                                                        *network_params)

        # creating and saving summary tables
        self.chapter7_summary = self._get_summary_table(
                                                Titles.chapter_7_directions)
        self.chapter8_summary = self._get_summary_table(
                                                Titles.chapter_8_directions)
        self.source_ch12_sum = self._get_summary_table(
                                                Titles.source_ch12_directions)
        self.network_ch12_sum = self._get_summary_table(
                                                Titles.network_ch12_directions)

        # self.wb, self.table_names = None, None

    def _get_wb_and_table_names(self, filename: str):
        wb = load_workbook(filename=filename, data_only=True)

        table_names = {}
        for worksheet in wb.sheetnames:
            for table in wb[worksheet].tables:
                table_names[table] = worksheet
        return wb, table_names

    def _get_table(self, table_name: str, *args) -> list[object]:
        def _keys_func(item: SourceEvent):
            x = str(item.event_years)[:4]
            y = str(item.event_years)[-4:]
            return int(x), int(y)

        worksheet = self.wb[self.table_names[table_name]]
        range = worksheet.tables[table_name].ref
        origin_data = worksheet[range][1:]  # bad
        import_data = []
        for row in origin_data:
            values = (cell.value for cell in row)
            import_data.append(self.classes[table_name](*values, *args))

        if table_name in (Titles.source_events, Titles.network_events):
            # Sorting
            import_data.sort(key=_keys_func)
            number = 0
            for item in import_data:
                number += 1
                item.number = number
            # Create a "Total" row
            values = (number + 1, Titles.total, import_data)
            return import_data, SummaryTable(*values)
        return import_data

    def _get_summary_table(self, direction_name: str,
                           events: SourceEvent = None):
        directions = self._get_table(direction_name)
        summary_tables = []
        events = (
            getattr(self, self.group_event[direction_name])
            if events is None else events
        )
        atribute_name = self.atributes[direction_name]
        for direct in directions:
            values = (direct.number, direct.name, events, atribute_name)
            summary_tables.append(SummaryTable(*values))
        return summary_tables

    def split_events_by_tso(self, type: str, attribute: str,
                            direction_name: str, cumulative: bool = False):
        if cumulative:
            sum_events_by_tso = defaultdict(list)
            sum_directions_by_tso = defaultdict(list)
            sum_totals = {}
        totals = {}
        number_by_tso = defaultdict(int)
        events_by_tso = defaultdict(list)
        directions_by_tso = defaultdict(list)
        # create events_by_tso
        for event in getattr(self, type + '_events'):
            tso = getattr(event, attribute)
            number_by_tso[tso] += 1
            event.number = number_by_tso[tso]
            events_by_tso[tso].append(event)
            if cumulative:
                sum_event = copy(event)
                sum_events_by_tso[tso].append(
                                        self._cum_total_events(sum_event))
        for tso, events in events_by_tso.items():
            # create directions_by_tso
            directions_by_tso[tso] = self._get_summary_table(
                                                    direction_name, events)
            # create totals
            values = (number_by_tso[tso] + 1, Titles.total, events)
            totals[tso] = SummaryTable(*values)
            if cumulative:
                sum_directions_by_tso[tso] = self._get_summary_table(
                    direction_name, sum_events_by_tso[tso]
                )
                values = (
                    number_by_tso[tso] + 1, Titles.total,
                    sum_events_by_tso[tso]
                )
                sum_totals[tso] = SummaryTable(*values)
        if cumulative:
            return (
                events_by_tso, totals, directions_by_tso,
                sum_events_by_tso, sum_totals,
                sum_directions_by_tso
            )
        return events_by_tso, totals, directions_by_tso

    def get_tso_name_list(self) -> dict:
        tso_list = self._get_table(Titles.tso_list)
        tso_name_list = {}
        for tso in tso_list:
            tso_name_list[tso.name] = tso.short_name
        return tso_name_list

    def _cum_total_events(self, sum_event: SummaryTable) -> SummaryTable:
        first_year = self.terms[0].start_year
        last_year = self.terms[0].end_year
        for year in range(first_year + 1, last_year + 1):
            setattr(
                sum_event, 'year_' + str(year),
                getattr(sum_event, 'year_' + str(year - 1)) +
                getattr(sum_event, 'year_' + str(year))
            )
        return sum_event
