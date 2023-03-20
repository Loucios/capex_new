import re
from dataclasses import fields

from base_datas import BaseEvent, Terms
from formats import (Chapter7Directions, Chapter8directions,
                     Chapter12Directions, NetworkTable, SheetNames,
                     SourceTable)
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from styles import Styles
from summary_datas import SummaryTable
from titles import Titles
from widths import Formats, Widths


class BaseTable:
    column = 1
    title_row = 1
    header_row = 3
    content_row = 5
    header_height = 70
    tables = {
        'source': SourceTable,
        'network': NetworkTable,
        'chapter_7_directions': Chapter7Directions,
        'chapter_8_directions': Chapter8directions,
        'network_ch12_directions': Chapter12Directions,
        'source_ch12_directions': Chapter12Directions,
    }
    styles = {
        Titles.base_style: Styles.style_1,
        Titles.title_style: Styles.style_4,
        Titles.header_style: Styles.style_3,
        Titles.footer_style: Styles.style_2,
        Titles.direction_style: Styles.style_5,
    }

    def __init__(self, workbook: Workbook, events: list[BaseEvent],
                 total: SummaryTable, terms: list[Terms],
                 tso: str | None = None, short_tso: str | None = None) -> None:
        self.wb = workbook
        self.events = events
        self.total = total
        self.start = terms[0].start_year
        self.end = terms[0].end_year
        self.tso = tso
        self.short_tso = short_tso
        self.table = self._get_table()
        self.title = self._get_title()
        self.sheet_name = self._get_sheet_name()
        self.merged_cells = []
        self.last_row = 0

        self._add_style()

    def _get_table(self):
        return self.tables[self.events[0].obj_type]

    def _get_names(self, type):
        value = getattr(SheetNames, self.events[0].obj_type)
        value = value[SheetNames.types[type]]
        return value

    def _get_title(self):
        middle = self._get_names('title_name')
        start = SheetNames.begining
        end = SheetNames.ending
        return f'{start}{middle}{end}'

    def _get_sheet_name(self):
        # return ''.join(w[0].upper() for w in self.name.split())
        return self._get_names('sheet_name')

    def _add_style(self) -> None:
        for title, style in self.styles.items():
            if title not in self.wb.named_styles:
                self.wb.add_named_style(style)

    def create_table(self) -> None:
        '''Create the table'''
        if self.sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(title=self.sheet_name)
            self._create_title()
            self._create_header()
        self._create_content()
        self._create_footer()
        self._merge_cells()

    def _create_title(self):
        '''Create title of our table'''
        self._set_cell_value(
            self.title_row, self.column, self.title, Titles.title_style,
            'title'
        )

    def _set_cell_value(self, row, column, value, style, title):
        ws = self.wb[self.sheet_name]
        ws.cell(row=row, column=column).value = value
        ws.cell(row=row, column=column).style = style
        form = getattr(Formats, title)
        ws.cell(row=row, column=column).number_format = form

    def _create_header(self):
        '''Create header of our table'''
        column = self.column
        row = self.header_row
        height = self.header_height
        style = Titles.header_style
        ws = self.wb[self.sheet_name]
        total = False
        year_column = 0

        for attribute in fields(self.table):
            title = attribute.name
            if 'year_' in title:
                if self.start <= int(title[-4:]) <= self.end:
                    ws.cell(row=row + 1, column=column).value = title[-4:]
                    year_column = column if not year_column else year_column
                    ws.cell(row=row, column=column).style = style
                    ws.cell(row=row + 1, column=column).style = style
                    self._set_column_width(title, column)
                    column += 1
            elif title == 'total':
                ws.cell(row=row + 1, column=column).value = getattr(
                                                                Titles, title)
                total = True
                ws.cell(row=row, column=column).style = style
                ws.cell(row=row + 1, column=column).style = style
                self._set_column_width(title, column)
                column += 1
            else:
                # merge first and second row
                ws.cell(row=row, column=column).style = style
                ws.cell(row=row + 1, column=column).style = style
                ws.merge_cells(
                    start_row=row, start_column=column,
                    end_row=row + 1, end_column=column
                )
                ws.cell(row=row, column=column).value = getattr(Titles, title)
                self._set_column_width(title, column)
                column += 1
        # merge first row under (years + total)
        period = self.end - self.start + (1 if total else 0)
        ws.merge_cells(
            start_row=row, start_column=year_column,
            end_row=row, end_column=year_column + period
        )
        ws.cell(row=row, column=year_column).value = Titles.capex_flow
        ws.row_dimensions[row + 1].height = height

    def _set_column_width(self, title, column):
        ws = self.wb[self.sheet_name]
        width = getattr(Widths, title)
        ws.column_dimensions[get_column_letter(column)].width = width

    def _create_content(self):
        '''Create event data rows of our table'''
        row = self.content_row
        for event in self.events:
            self._set_row_data(event, row, Titles.base_style)
            row += 1
        self.last_row = row

    def _set_row_data(self, obj, row, style):
        column = self.column
        for attribute in fields(self.table):
            title = attribute.name
            style = style
            # Create table between start and end year
            if ('year_' not in title or (
                    self.start <= int(title[-4:]) <= self.end
               )):
                value = self._get_value(row, obj, title)
                self._set_cell_value(row, column, value, style, title)
                # Merge cells around direction title
                self._set_merged_cells(row, column, obj, value, title)
                column += 1
        # Set addition height with direction row
        self._set_height(row, obj)

    def _get_value(self, row, obj, title):
        if title == 'number':
            return row - self.content_row + 1
        return getattr(obj, title, '')

    def _set_merged_cells(self, row, column, obj, value, title):
        if title == 'event_title' and isinstance(obj, SummaryTable):
            self.merged_cells.append(
                {'row': row, 'column': column,
                 'value': value, 'title': title}
            )

    def _set_height(self, row, obj):
        if isinstance(obj, SummaryTable):
            ws = self.wb[self.sheet_name]
            ws.row_dimensions[row].height = self.header_height

    def _create_footer(self):
        self._set_row_data(
            self.total, self.last_row, Titles.footer_style
        )

    def _merge_cells(self):
        ws = self.wb[self.sheet_name]
        for cells in self.merged_cells:
            ws.merge_cells(
                start_row=cells['row'],
                start_column=self.column + 1,
                end_row=cells['row'],
                end_column=cells['column']
            )
            self._set_cell_value(
                cells['row'], self.column + 1, cells['value'],
                Titles.footer_style, cells['title'])


class DirectionsTable(BaseTable):

    def __init__(self, workbook, events, total, terms,
                 sums: list[SummaryTable], tso=None, short_tso=None) -> None:
        self.sums = sums
        super().__init__(workbook, events, total, terms, tso, short_tso)

        self._sort_events()

    def _get_table(self):
        return self.tables[self.sums[0].directions]

    def _get_names(self, type):
        value = getattr(SheetNames, self.sums[0].directions)
        value = value[SheetNames.types[type]]
        return value

    def _sort_events(self):
        directions = self.sums[0].directions
        self.events.sort(
            key=lambda x: getattr(x, directions), reverse=False
        )

    def _create_content(self):
        super()._create_content()
        self._create_directions()

    def _create_directions(self) -> None:
        '''Insert to existing table "Summary by directions" rows'''
        self.last_row = 0
        row = self.content_row
        self.merged_cells = []
        for direction in self.sums:
            self.wb[self.sheet_name].insert_rows(row)
            self._set_row_data(direction, row, Titles.footer_style)
            row += (direction.amount + 1)
        self.last_row = row

    def _get_value(self, row, obj, title):
        if title != 'number':
            return super()._get_value(row, obj, title)
        return self._get_number(row, obj)

    def _get_number(self, row, obj):
        if isinstance(obj, SummaryTable):
            if row == self.last_row:
                return len(self.sums) + 1
            return obj.number
        value = row - self.content_row + 1
        left = 0
        for index, direction in enumerate(self.sums):
            if index:
                left += self.sums[index - 1].amount
            right = left + direction.amount
            if left < value <= right:
                number = value - left
                return f'{direction.number}.{number}'
        return value


class ByTSOTable(DirectionsTable):

    def _get_sheet_name(self):
        sheet_name = super()._get_sheet_name()
        short_name = self._get_short_name()
        return f'{sheet_name} {short_name}'

    def _get_short_name(self):
        if self.short_tso is None:
            pattern = re.compile(r'[^\w ]+')
            return re.sub(pattern, '', self.tso)[:10]
        return self.short_tso

    def _get_title(self):
        title = super()._get_title()
        middle = SheetNames.tso_begining
        return f'{title}{middle}{self.tso}'


class DirectionsTableSum(ByTSOTable):
    content_row = 4

    def __init__(self, workbook, events, total, terms,
                 sums, sum_events, sum_total, sum_sums,
                 tso=None, short_tso=None) -> None:
        self.sum_events = sum_events
        self.sum_sums = sum_sums
        self.sum_total = sum_total
        super().__init__(workbook, events, total, terms,
                         sums, tso, short_tso)

    def _sort_events(self):
        super()._sort_events()
        directions = self.sum_sums[0].directions
        self.sum_events.sort(
            key=lambda x: getattr(x, directions), reverse=False
        )

    def _create_footer(self):
        row = self.content_row
        self._set_row_data(self.total, row, Titles.footer_style)
        self.sum_total.row_name = Titles.row_name_second
        self._set_row_data(self.sum_total, row + 1, Titles.footer_style)
        row += 2
        self.last_row = row

    def _create_content(self):
        row = self.last_row
        for event, sum_event in zip(self.events, self.sum_events):
            self._set_row_data(event, row, Titles.base_style)
            sum_event.row_name = Titles.row_name_second
            self._set_row_data(sum_event, row + 1, Titles.base_style)
            row += 2
        row = self.last_row
        for direction, sum_direction in zip(self.sums, self.sum_sums):
            self.wb[self.sheet_name].insert_rows(row)
            self._set_row_data(direction, row, Titles.base_style)
            self.wb[self.sheet_name].insert_rows(row + 1)
            sum_direction.row_name = Titles.row_name_second
            self._set_row_data(sum_direction, row + 1, Titles.base_style)
            row += (direction.amount + 1) * 2
        self.last_row = row

    def create_table(self) -> None:
        '''Create the table'''
        if self.sheet_name not in self.wb.sheetnames:
            self.wb.create_sheet(title=self.sheet_name)
            self._create_title()
            self._create_header()
        self._create_footer()
        self._create_content()

    def _create_header(self):
        '''Create header of our table'''
        column = self.column
        row = self.header_row
        # height = self.header_height
        style = Titles.header_style
        ws = self.wb[self.sheet_name]

        for attribute in fields(self.table):
            title = attribute.name
            if 'year_' in title:
                if self.start <= int(title[-4:]) <= self.end:
                    ws.cell(row=row, column=column).value = title[-4:]
                    ws.cell(row=row, column=column).style = style
                    self._set_column_width(title, column)
                    column += 1
            elif title == 'total':
                ws.cell(row=row, column=column).value = getattr(Titles, title)
                ws.cell(row=row, column=column).style = style
                self._set_column_width(title, column)
                column += 1
            else:
                # merge first and second row
                ws.cell(row=row, column=column).style = style
                ws.cell(row=row, column=column).value = getattr(Titles, title)
                self._set_column_width(title, column)
                column += 1
        # ws.row_dimensions[row + 1].height = height
